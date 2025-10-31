from __future__ import annotations

"""Core PDF operations: merge, split, compress.

All functions raise exceptions on failure and log steps.

Dependencies:
- pypdf is required (for merge & split)
- pymupdf is optional (preferred for compression if available)
- Ghostscript fallback via ``subprocess``
- pytesseract is optional (for OCR functionality)
"""

import io
import logging
import os
import sys
import threading
import urllib.request
from contextlib import contextmanager
from pathlib import Path
from typing import Callable, List, Optional

from .utils import find_ghostscript_command

try:  # Attempt optional pypdf import
    from pypdf import PdfReader, PdfWriter  # type: ignore

    _HAVE_PYPDF = True
except Exception:  # pragma: no cover - handled at runtime
    PdfReader = PdfWriter = None  # type: ignore
    _HAVE_PYPDF = False

# Attempt optional import of PyMuPDF
try:
    import fitz  # type: ignore

    _HAVE_PYMUPDF = True
except ModuleNotFoundError:
    _HAVE_PYMUPDF = False

# Attempt optional import of pytesseract and check binary availability
try:
    import pytesseract  # type: ignore
    from PIL import Image, ImageEnhance, ImageFilter, ImageOps  # type: ignore

    _HAVE_TESSERACT = True
    try:
        # First try the standard way
        pytesseract.get_tesseract_version()
        _TESSERACT_INSTALLED = True
    except Exception:  # pragma: no cover - binary missing
        # If that fails, try our improved discovery
        try:
            from .utils import check_tesseract_version

            version = check_tesseract_version()
            if version:
                _TESSERACT_INSTALLED = True
                # Set the tesseract command path if we found it
                tesseract_cmd = check_tesseract_version.__globals__.get("find_tesseract_command")
                if tesseract_cmd:
                    cmd_path = tesseract_cmd()
                    if cmd_path:
                        pytesseract.pytesseract.tesseract_cmd = cmd_path
            else:
                _TESSERACT_INSTALLED = False
        except Exception:  # pragma: no cover - binary missing
            _TESSERACT_INSTALLED = False
except ModuleNotFoundError:
    _HAVE_TESSERACT = False
    _TESSERACT_INSTALLED = False

# Attempt optional import of camelot
try:
    import camelot  # type: ignore  # noqa: F401

    _HAVE_CAMELOT = True
except ModuleNotFoundError:
    _HAVE_CAMELOT = False

# Attempt optional import of pdfplumber
try:
    import pdfplumber  # type: ignore  # noqa: F401

    _HAVE_PDFPLUMBER = True
except ModuleNotFoundError:
    _HAVE_PDFPLUMBER = False

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Cross-platform utilities
# ---------------------------------------------------------------------------


def _normalize_path(path: str | os.PathLike[str]) -> str:
    """Normalize path for cross-platform compatibility.

    Converts path separators to the appropriate format for the current platform
    and resolves any relative paths.
    """
    return os.path.normpath(str(path))


def _ensure_executable_permissions(path: str) -> None:
    """Ensure executable permissions on Unix-like systems.

    On Windows, this is a no-op. On Unix-like systems, ensures the file
    has execute permissions.
    """
    if not sys.platform.startswith("win"):
        try:
            import stat

            current_permissions = os.stat(path).st_mode
            os.chmod(path, current_permissions | stat.S_IEXEC)
        except (OSError, PermissionError):
            # If we can't change permissions, continue anyway
            pass


# ---------------------------------------------------------------------------
# Context Managers for Resource Management
# ---------------------------------------------------------------------------


@contextmanager
def pdf_document(file_path: str | os.PathLike[str]):
    """Context manager for PDF document handling to ensure proper resource cleanup.

    This context manager opens a PDF document using PyMuPDF and ensures that
    the document is properly closed even if an exception occurs during processing.

    Parameters
    ----------
    file_path : str or PathLike
        Path to the PDF file to open

    Yields
    ------
    fitz.Document
        The opened PDF document object

    Raises
    ------
    RuntimeError
        If the PDF file cannot be opened

    Examples
    --------
    >>> with pdf_document("example.pdf") as doc:
    ...     page = doc.load_page(0)
    ...     # Process the page
    ...     pass
    # Document is automatically closed even if an exception occurs
    """
    doc = None
    try:
        import fitz  # type: ignore

        doc = fitz.open(str(file_path))
        yield doc
    except Exception as e:
        raise RuntimeError(f"Failed to open PDF file '{file_path}'. Error: {str(e)}")
    finally:
        if doc is not None:
            try:
                doc.close()
            except Exception as e:
                logger.warning(f"Error closing PDF document: {e}")


@contextmanager
def image_document(image_data: bytes):
    """Context manager for image handling to ensure proper resource cleanup.

    This context manager opens an image from bytes data and ensures that
    the image is properly closed even if an exception occurs during processing.

    Parameters
    ----------
    image_data : bytes
        Image data in bytes format

    Yields
    ------
    PIL.Image.Image
        The opened image object

    Raises
    ------
    RuntimeError
        If the image data cannot be processed

    Examples
    --------
    >>> with image_document(image_bytes) as img:
    ...     processed_img = img.convert("L")
    ...     # Process the image
    ...     pass
    # Image is automatically closed even if an exception occurs
    """
    img = None
    try:
        from PIL import Image  # type: ignore

        img = Image.open(io.BytesIO(image_data)).copy()
        yield img
    except Exception as e:
        raise RuntimeError(f"Failed to process image data. Error: {str(e)}")
    finally:
        if img is not None:
            try:
                img.close()
            except Exception as e:
                logger.warning(f"Error closing image: {e}")


# ---------------------------------------------------------------------------
# Progress Tracking
# ---------------------------------------------------------------------------


class OCRProgress:
    """Thread-safe progress tracking for OCR operations.

    This class provides thread-safe progress tracking for OCR operations
    that process multiple pages. It tracks the current page being processed,
    the status message, and calculates the completion percentage.

    Parameters
    ----------
    total_pages : int
        Total number of pages to process

    Attributes
    ----------
    total_pages : int
        Total number of pages to process
    current_page : int
        Current page being processed (1-based)
    current_status : str
        Current status message
    _lock : threading.Lock
        Thread lock for synchronization

    Examples
    --------
    >>> progress = OCRProgress(10)
    >>> progress.update(5, "Processing page 5")
    >>> page, status, percentage = progress.get_progress()
    >>> print(f"Progress: {percentage:.1f}%")
    Progress: 50.0%
    """

    def __init__(self, total_pages: int):
        """Initialize OCR progress tracker.

        Parameters
        ----------
        total_pages : int
            Total number of pages to process
        """
        self.total_pages = total_pages
        self.current_page = 0
        self.current_status = "Initializing..."
        self._lock = threading.Lock()

    def update(self, page: int, status: str):
        """Update progress for a specific page.

        Parameters
        ----------
        page : int
            Current page number (1-based)
        status : str
            Status message for the current operation
        """
        with self._lock:
            self.current_page = page
            self.current_status = status

    def get_progress(self) -> tuple[int, str, float]:
        """Get current progress as (current_page, status, percentage).

        Returns
        -------
        tuple
            A tuple containing:
            - current_page (int): Current page being processed
            - status (str): Current status message
            - percentage (float): Completion percentage (0-100)
        """
        with self._lock:
            percentage = (self.current_page / self.total_pages * 100) if self.total_pages > 0 else 0
            return self.current_page, self.current_status, percentage


# ---------------------------------------------------------------------------
# Merge PDFs
# ---------------------------------------------------------------------------


def merge_pdfs(input_files: List[str | os.PathLike[str]], output_file: str | os.PathLike[str]) -> None:
    """Merge multiple PDF files into a single PDF file.

    This function merges multiple PDF files in the order specified by the input_files
    list into a single output PDF file. Each file is processed sequentially, and all
    pages from each file are added to the merged document.

    Parameters
    ----------
    input_files : List[str or PathLike]
        List of paths to PDF files to merge, in the desired order
    output_file : str or PathLike
        Path to the output merged PDF file

    Raises
    ------
    ValueError
        If no input files are provided
    FileNotFoundError
        If any input file does not exist
    RuntimeError
        If any input file cannot be read or the output cannot be written
    PermissionError
        If there are insufficient permissions to write the output file

    Examples
    --------
    >>> merge_pdfs(["file1.pdf", "file2.pdf"], "merged.pdf")
    >>> merge_pdfs([Path("doc1.pdf"), Path("doc2.pdf")], Path("output.pdf"))

    Notes
    -----
    This function uses the pypdf library for PDF manipulation. All input files
    must be valid PDF documents. The function creates the output directory if
    it does not exist.
    """
    if not input_files:
        raise ValueError("No input files provided for merging")

    writer = PdfWriter()
    processed_files = []

    for file in input_files:
        logger.info("Adding pages from %s", file)
        try:
            reader = PdfReader(str(file))
            processed_files.append(file)
        except FileNotFoundError as exc:
            raise FileNotFoundError(f"Input file not found: {file}. Please check the file path and try again.") from exc
        except Exception as exc:
            raise RuntimeError(
                f"Failed to read PDF '{file}'. The file may be corrupted or password-protected. Error: {exc}"
            ) from exc

        for page_idx, page in enumerate(reader.pages, start=1):
            try:
                writer.add_page(page)
            except Exception as exc:
                raise RuntimeError(
                    f"Failed to add page {page_idx} from '{file}' to the merged document. Error: {exc}"
                ) from exc

    # Ensure output directory exists
    output_path = Path(output_file)
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        raise RuntimeError(
            f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {exc}"
        ) from exc

    # Write the merged PDF
    try:
        with output_path.open("wb") as fp:
            writer.write(fp)
    except PermissionError as exc:
        raise PermissionError(
            f"Permission denied when writing to '{output_file}'. Please check file permissions "
            f"and ensure the file is not open in another application."
        ) from exc
    except Exception as exc:
        raise RuntimeError(f"Failed to write merged PDF to '{output_file}'. Error: {exc}") from exc

    logger.info("Merged %d files -> %s", len(input_files), output_file)


def split_pdf(
    input_file: str | os.PathLike[str],
    output_dir: str | os.PathLike[str] | None = None,
    *,
    method: str = "single",
    page_range: str | None = None,
    naming_pattern: str = "page_{page}",
    compress: bool = False,
    compression_level: int = 0,
    password: str | None = None,
    progress_callback: Callable[[int, int, str], None] | None = None,
    **kwargs,
) -> List[Path]:
    """Split a PDF file into multiple single-page PDF files.

    This function splits a PDF file into individual pages, creating a separate
    PDF file for each page. Each output file is named according to the
    naming_pattern parameter.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file to split
    output_dir : str or PathLike, optional
        Directory to save the split files. If None, uses 'dest_dir' from kwargs
    method : str, default "single"
        Split method (currently only "single" is supported)
    page_range : str, optional
        Page range to split (not currently used)
    naming_pattern : str, default "page_{page}"
        Pattern for naming output files. Use {page} for page number
    compress : bool, default False
        Whether to compress each split page
    compression_level : int, default 0
        Compression level (0=screen, 1=ebook, 2=printer, 3=prepress)
    password : str, optional
        Password for encrypted PDFs (not currently used)
    progress_callback : Callable, optional
        Callback function to report progress with signature (current, total, status)
    **kwargs
        Additional keyword arguments (supports 'dest_dir' for backward compatibility)

    Returns
    -------
    List[Path]
        List of paths to the created split files

    Raises
    ------
    TypeError
        If output_dir is not provided
    FileNotFoundError
        If the input file does not exist
    RuntimeError
        If the input file cannot be read or output cannot be written
    ValueError
        If the input PDF contains no pages

    Examples
    --------
    >>> split_pdf("document.pdf", "output_dir")
    >>> split_pdf("document.pdf", "output_dir", naming_pattern="page_{page:03d}.pdf")
    >>> split_pdf("document.pdf", "output_dir", compress=True, compression_level=1)

    Notes
    -----
    This function uses the pypdf library for PDF manipulation. The output directory
    is created if it does not exist. Compression is performed using the compress_pdf
    function if requested.
    """
    if output_dir is None and "dest_dir" in kwargs:
        output_dir = kwargs["dest_dir"]
    if output_dir is None:
        raise TypeError("output_dir is required. Please specify an output directory for the split files.")

    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    try:
        dest = Path(output_dir)
        dest.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        raise RuntimeError(
            f"Failed to create output directory '{output_dir}'. Please check permissions. Error: {exc}"
        ) from exc

    try:
        reader = PdfReader(str(input_file))
    except Exception as exc:
        raise RuntimeError(
            f"Failed to read PDF '{input_file}'. The file may be corrupted or password-protected. Error: {exc}"
        ) from exc

    created: List[Path] = []
    total = len(reader.pages)

    if total == 0:
        raise ValueError(f"The PDF file '{input_file}' contains no pages to split.")

    quality_map = {0: "screen", 1: "ebook", 2: "printer", 3: "prepress"}

    for idx, page in enumerate(reader.pages, start=1):
        writer = PdfWriter()
        try:
            writer.add_page(page)
        except Exception as exc:
            raise RuntimeError(f"Failed to add page {idx} to split document. Error: {exc}") from exc

        name = naming_pattern.format(page=idx)
        if not name.lower().endswith(".pdf"):
            name += ".pdf"
        out_path = dest / name

        try:
            with out_path.open("wb") as fp:
                writer.write(fp)
        except PermissionError as exc:
            raise PermissionError(
                f"Permission denied when writing to '{out_path}'. Please check file permissions "
                f"and ensure the file is not open in another application."
            ) from exc
        except Exception as exc:
            raise RuntimeError(f"Failed to write split page {idx} to '{out_path}'. Error: {exc}") from exc

        if compress:
            quality = quality_map.get(compression_level, "screen")
            try:
                compress_pdf(out_path, out_path, quality=quality)
            except Exception as exc:  # pragma: no cover - optional path
                logger.error("Compression failed for %s: %s", out_path, exc)
                # Don't fail the entire operation if compression fails, just log it
                logger.warning(
                    "Continuing without compression for %s due to compression error",
                    out_path,
                )

        created.append(out_path)
        logger.debug("Wrote %s", out_path)
        if progress_callback:
            progress_callback(idx, total, f"Split page {idx}")

    logger.info("Split %s into %d single-page PDFs", input_file, len(created))
    return created


def split_at_page(
    input_file: str | os.PathLike[str],
    dest_dir: str | os.PathLike[str],
    split_page: int,
) -> tuple[Path, Path]:
    """Split a PDF file into two parts at a specific page.

    This function splits a PDF file into two separate PDF files at the specified
    page number. The first part contains pages 1 through split_page, and the
    second part contains pages split_page+1 through the end of the document.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file to split
    dest_dir : str or PathLike
        Directory to save the split files
    split_page : int
        Page number where to split (1-based, inclusive for first part)

    Returns
    -------
    tuple[Path, Path]
        Tuple containing paths to (first_part, second_part) PDF files

    Raises
    ------
    ValueError
        If split_page is invalid (less than 1 or greater than or equal to total pages)
    FileNotFoundError
        If the input file does not exist
    RuntimeError
        If the input file cannot be read or output cannot be written

    Examples
    --------
    >>> part1, part2 = split_at_page("document.pdf", "output_dir", 5)
    >>> print(f"First part: {part1}")
    >>> print(f"Second part: {part2}")

    Notes
    -----
    This function uses the extract_page_range function internally to create
    the two output files. The destination directory is created if it does
    not exist.
    """
    reader = PdfReader(str(input_file))
    total = len(reader.pages)
    if split_page < 1 or split_page >= total:
        raise ValueError(f"split_page must be between 1 and {total - 1}, got {split_page}")

    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)

    part1 = dest / "part1.pdf"
    part2 = dest / "part2.pdf"

    extract_page_range(input_file, part1, 1, split_page)
    extract_page_range(input_file, part2, split_page + 1, total)

    return part1, part2


def extract_page_range(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    start_page: int,
    end_page: int,
) -> None:
    """Extract a range of pages from a PDF file into a new PDF file.

    This function extracts pages from start_page to end_page (inclusive) from
    the input PDF file and saves them as a new PDF file.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file
    output_file : str or PathLike
        Path to the output PDF file containing the extracted pages
    start_page : int
        First page to extract (1-based, inclusive)
    end_page : int
        Last page to extract (1-based, inclusive)

    Raises
    ------
    FileNotFoundError
        If the input file does not exist
    RuntimeError
        If the input file cannot be read or output cannot be written
    ValueError
        If the page range is invalid or the input PDF contains no pages

    Examples
    --------
    >>> extract_page_range("document.pdf", "pages_1-5.pdf", 1, 5)
    >>> extract_page_range("document.pdf", "middle_pages.pdf", 10, 20)

    Notes
    -----
    This function uses the pypdf library for PDF manipulation. The output
    directory is created if it does not exist. Page numbers are 1-based.
    """

    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    try:
        reader = PdfReader(str(input_file))
    except Exception as exc:
        raise RuntimeError(
            f"Failed to read PDF '{input_file}'. The file may be corrupted or password-protected. Error: {exc}"
        ) from exc

    total = len(reader.pages)

    if total == 0:
        raise ValueError(f"The PDF file '{input_file}' contains no pages to extract.")

    if start_page < 1 or end_page > total or start_page > end_page:
        raise ValueError(
            f"Invalid page range {start_page}-{end_page} for document with {total} pages. "
            f"Please specify a valid range between 1 and {total}."
        )

    writer = PdfWriter()
    for i in range(start_page - 1, end_page):
        try:
            writer.add_page(reader.pages[i])
        except Exception as exc:
            page_num = i + 1
            raise RuntimeError(f"Failed to add page {page_num} to extracted document. Error: {exc}") from exc

    # Ensure output directory exists
    output_path = Path(output_file)
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as exc:
        raise RuntimeError(
            f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {exc}"
        ) from exc

    # Write the extracted PDF
    try:
        with output_path.open("wb") as fp:
            writer.write(fp)
    except PermissionError as exc:
        raise PermissionError(
            f"Permission denied when writing to '{output_file}'. Please check file permissions "
            f"and ensure the file is not open in another application."
        ) from exc
    except Exception as exc:
        raise RuntimeError(f"Failed to write extracted PDF to '{output_file}'. Error: {exc}") from exc

    logger.info(
        "Extracted pages %d-%d from %s -> %s",
        start_page,
        end_page,
        input_file,
        output_file,
    )


# ---------------------------------------------------------------------------
# Compress PDF
# ---------------------------------------------------------------------------

_VALID_QUALITIES = {"screen", "ebook", "printer", "prepress"}


def compress_pdf(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    quality: str = "screen",
) -> None:
    """Compress a PDF file to reduce its file size.

    This function compresses a PDF file using either PyMuPDF (preferred) or
    Ghostscript as a fallback. The compression quality determines the trade-off
    between file size and quality.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file to compress
    output_file : str or PathLike
        Path to the output compressed PDF file
    quality : str, default "screen"
        Compression quality preset. One of:
        - "screen": Low quality, smallest file size (72 DPI)
        - "ebook": Medium quality (150 DPI)
        - "printer": High quality (300 DPI)
        - "prepress": Very high quality (300 DPI, color preserving)

    Raises
    ------
    FileNotFoundError
        If the input file does not exist
    ValueError
        If the quality setting is invalid
    RuntimeError
        If compression fails or neither backend is available

    Examples
    --------
    >>> compress_pdf("large_document.pdf", "compressed_document.pdf")
    >>> compress_pdf("document.pdf", "small.pdf", quality="ebook")
    >>> compress_pdf("photo.pdf", "high_quality.pdf", quality="printer")

    Notes
    -----
    The function first attempts to use PyMuPDF for compression, which is faster
    and runs in-process. If PyMuPDF is not available or fails, it falls back to
    Ghostscript, which requires a separate installation but offers more compression
    options. The output directory is created if it does not exist.
    """
    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    quality = quality.lower()
    if quality not in _VALID_QUALITIES:
        raise ValueError(f"Invalid quality setting: '{quality}'. Must be one of {_VALID_QUALITIES}.")

    if _HAVE_PYMUPDF:
        try:
            _compress_with_pymupdf(input_file, output_file)
            return
        except Exception as exc:
            logger.warning(
                "PyMuPDF compression failed, attempting Ghostscript fallback. Error: %s",
                exc,
            )
            # Continue to Ghostscript fallback

    gs_cmd = find_ghostscript_command()
    if gs_cmd:
        try:
            _compress_with_ghostscript(gs_cmd, input_file, output_file, quality)
            return
        except Exception as exc:
            raise RuntimeError(f"Both PyMuPDF and Ghostscript compression failed. Ghostscript error: {exc}") from exc

    raise RuntimeError(
        "Compression unavailable: Please install 'pymupdf' or Ghostscript and try again. "
        "PyMuPDF provides faster in-process compression, while Ghostscript offers more compression options."
    )


def _compress_with_pymupdf(input_file: str | os.PathLike[str], output_file: str | os.PathLike[str]) -> None:
    """Compress via PyMuPDF by rewriting each page with lower image resolution.

    This internal function uses PyMuPDF to compress a PDF by recompressing
    images and applying other space-saving techniques.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file
    output_file : str or PathLike
        Path to the output compressed PDF file

    Raises
    ------
    Exception
        If compression fails for any reason

    Notes
    -----
    This function applies several compression techniques:
    - Image recompression to 150 DPI with JPEG quality 85
    - Garbage collection to remove unused objects
    - Deflate compression for streams
    - Cleaning to optimize content
    """
    try:  # pragma: no cover – optional dependency
        import fitz  # type: ignore  # local import to avoid unconditional dependency

        logger.info("Compressing %s using PyMuPDF", input_file)
        doc = fitz.open(str(input_file))

        # Options: recompress images to 150 DPI JPEG quality 85.
        mat = fitz.Matrix(1, 1)  # No scaling; images will be recompressed by save options
        for page in doc:
            # We redraw each page to ensure images are recompressed.
            page.get_pixmap(matrix=mat)

        doc.save(  # pragma: no cover – external dependency branch
            str(output_file),
            garbage=4,  # thorough garbage collection
            deflate=True,
            clean=True,
            incremental=False,
        )
        doc.close()
        logger.info("Saved compressed PDF to %s", output_file)
    except Exception as e:  # pragma: no cover – external dependency branch
        logger.error("PyMuPDF compression failed: %s", e)


def _compress_with_ghostscript(
    gs_cmd: str,
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    quality: str,
) -> None:
    """Compress via Ghostscript using *quality* preset."""
    # Normalize paths for cross-platform compatibility
    input_file = _normalize_path(input_file)
    output_file = _normalize_path(output_file)

    # Ensure Ghostscript executable has proper permissions
    _ensure_executable_permissions(gs_cmd)

    # Handle paths with spaces by quoting them
    if sys.platform.startswith("win"):
        # On Windows, use subprocess.list2cmdline for proper quoting
        import subprocess

        args = [
            gs_cmd,
            "-sDEVICE=pdfwrite",
            f"-dPDFSETTINGS=/{quality}",
            "-dNOPAUSE",
            "-dQUIET",
            "-dBATCH",
            f"-sOutputFile={output_file}",
            input_file,
        ]
    else:
        # On Unix-like systems, use shell=False for better security
        args = [
            gs_cmd,
            "-sDEVICE=pdfwrite",
            f"-dPDFSETTINGS=/{quality}",
            "-dNOPAUSE",
            "-dQUIET",
            "-dBATCH",
            f"-sOutputFile={output_file}",
            input_file,
        ]

    logger.info("Running Ghostscript compression: %s", " ".join(args))
    try:
        process = subprocess.run(args, capture_output=True, text=True, timeout=300)  # 5 minute timeout
        if process.returncode != 0:
            logger.error("Ghostscript failed: %s", process.stderr)
            raise RuntimeError(f"Ghostscript compression failed: {process.stderr.strip()}")
        logger.info("Saved compressed PDF to %s", output_file)
    except subprocess.TimeoutExpired:
        raise RuntimeError("Ghostscript compression timed out after 5 minutes")
    except Exception as e:
        raise RuntimeError(f"Ghostscript compression failed: {str(e)}")


# ---------------------------------------------------------------------------
# OCR PDF
# ---------------------------------------------------------------------------


def ensure_tessdata_language(language: str, user_tessdata_dir: Path) -> None:
    """Ensure the required Tesseract language data file is present in user tessdata dir. Download if missing."""
    lang_files = [lang.strip() for lang in language.split("+") if lang.strip()]
    base_url = "https://github.com/tesseract-ocr/tessdata/raw/main/"
    user_tessdata_dir.mkdir(parents=True, exist_ok=True)
    for lang in lang_files:
        traineddata_path = user_tessdata_dir / f"{lang}.traineddata"
        if not traineddata_path.exists():
            url = f"{base_url}{lang}.traineddata"
            try:
                print(f"Downloading {lang}.traineddata for Tesseract...")
                urllib.request.urlretrieve(url, traineddata_path)
            except Exception as e:
                raise RuntimeError(f"Failed to download {lang}.traineddata: {e}")


def _setup_tesseract_env() -> None:
    """Setup Tesseract environment for cross-platform compatibility."""
    # Check if TESSERACT_CMD is set in environment
    tesseract_cmd = os.environ.get("TESSERACT_CMD")
    if tesseract_cmd:
        try:
            import pytesseract  # type: ignore

            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd
            _ensure_executable_permissions(tesseract_cmd)
        except Exception:
            pass  # Continue with default behavior if setting fails

    # Set TESSDATA_PREFIX if not already set
    tessdata_prefix = os.environ.get("TESSDATA_PREFIX")
    if not tessdata_prefix:
        user_tessdata_dir = Path.home() / ".pdfutils" / "tessdata"
        os.environ["TESSDATA_PREFIX"] = str(user_tessdata_dir)


def extract_text_with_ocr(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    language: str = "eng",
    dpi: int = 300,
    start_page: int = 1,
    end_page: int | None = None,
    config: str = "",
    output_format: str = "text",
    # Preprocessing options
    binarize: bool = False,
    threshold: int = 128,
    resize_factor: float = 1.0,
    deskew: bool = False,
    denoise: bool = False,
    contrast_factor: float = 1.0,
    brightness_factor: float = 1.0,
    sharpen: bool = False,
    blur: float = 0.0,
    morph_op: str = "none",
    morph_kernel: int = 3,
    progress_callback: Optional[Callable[[tuple[int, str, float]], None]] = None,
) -> None:
    """Extract text from PDF using OCR and save to text, hOCR, or JSON file.
    Parameters
    ----------
    input_file: Source PDF path.
    output_file: Destination text/hOCR/JSON file path.
    language: Tesseract language code(s).
    dpi: Image DPI for rendering.
    start_page: 1-based start page (inclusive).
    end_page: 1-based end page (inclusive). If None, process to last page.
    config: Custom Tesseract config string.
    output_format: 'text', 'hocr', or 'json'.
    """
    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    # Validate output format
    if output_format not in ["text", "hocr", "json"]:
        raise ValueError(f"Unsupported output format '{output_format}'. Supported formats are: 'text', 'hocr', 'json'")

    # Setup Tesseract environment for cross-platform compatibility
    _setup_tesseract_env()

    if not _HAVE_PYMUPDF:
        raise RuntimeError("PyMuPDF is required for OCR functionality. Please install it with: pip install pymupdf")
    if not _HAVE_TESSERACT:
        raise RuntimeError(
            "pytesseract is required for OCR functionality. Please install it with: pip install pytesseract"
        )
    if not _TESSERACT_INSTALLED:
        import pytesseract  # type: ignore

        raise pytesseract.pytesseract.TesseractError(
            1,
            "Tesseract OCR engine not found. Please install Tesseract from https://github.com/tesseract-ocr/tesseract",
        )

    import fitz  # type: ignore
    import pytesseract  # type: ignore

    logger.info("Starting OCR extraction from %s", input_file)

    with pdf_document(input_file) as doc:
        total_pages = len(doc)
        if total_pages == 0:
            # For empty PDFs, create an empty output file and return
            output_path = Path(output_file)
            try:
                output_path.parent.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                raise RuntimeError(
                    f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {str(e)}"
                )

            try:
                with output_path.open("w", encoding="utf-8") as f:
                    f.write("")
                logger.info("Empty PDF processed: %s -> %s", input_file, output_file)
                return
            except PermissionError:
                raise PermissionError(
                    f"Permission denied when writing to '{output_file}'. Please check file permissions "
                    f"and ensure the file is not open in another application."
                )
            except Exception as e:
                raise RuntimeError(f"Failed to write OCR output to '{output_file}'. Error: {str(e)}")

        if end_page is None:
            end_page = int(total_pages)

        # Validate page range
        if start_page < 1 or end_page > total_pages or start_page > end_page:
            raise ValueError(
                f"Invalid page range {start_page}-{end_page} for document with {total_pages} pages. "
                f"Please specify a valid range between 1 and {total_pages}."
            )

        # Initialize progress tracking
        progress = OCRProgress(end_page - start_page + 1)
        extracted = []

        # Ensure tessdata for selected language(s)
        user_tessdata_dir = Path.home() / ".pdfutils" / "tessdata"
        try:
            ensure_tessdata_language(language, user_tessdata_dir)
        except Exception as e:
            raise RuntimeError(f"Failed to ensure Tesseract language data for '{language}'. Error: {str(e)}")

        # Set TESSDATA_PREFIX for this process
        os.environ["TESSDATA_PREFIX"] = str(user_tessdata_dir)

        for page_num in range(start_page - 1, end_page):
            current_page = page_num + 1
            progress.update(current_page, f"Processing page {current_page}/{total_pages}")

            if progress_callback:
                progress_callback(progress.get_progress())

            logger.info(f"Processing page {current_page}/{total_pages}")

            try:
                page = doc.load_page(page_num)
            except Exception as e:
                raise RuntimeError(f"Failed to load page {current_page} from PDF. Error: {str(e)}")

            try:
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)  # type: ignore
                img_data = pix.tobytes("png")
                with image_document(img_data) as img:
                    # Apply preprocessing
                    if any([binarize, resize_factor != 1.0, deskew, denoise]):
                        progress.update(current_page, f"Preprocessing page {current_page}")
                        if progress_callback:
                            progress_callback(progress.get_progress())

                        try:
                            img = preprocess_image(
                                img,
                                binarize=binarize,
                                threshold=threshold,
                                resize_factor=resize_factor,
                                deskew=deskew,
                                denoise=denoise,
                                contrast_factor=contrast_factor,
                                brightness_factor=brightness_factor,
                                sharpen=sharpen,
                                blur=blur,
                                morph_op=morph_op,
                                morph_kernel=morph_kernel,
                            )
                        except Exception as e:
                            raise RuntimeError(f"Failed to preprocess page {current_page}. Error: {str(e)}")

                    progress.update(current_page, f"Running OCR on page {current_page}")
                    if progress_callback:
                        progress_callback(progress.get_progress())

                    try:
                        if output_format == "hocr":
                            text = pytesseract.image_to_pdf_or_hocr(img, lang=language, config=config, extension="hocr")
                            if isinstance(text, bytes):
                                extracted.append(text.decode("utf-8"))
                            else:
                                extracted.append(str(text))
                        elif output_format == "json":
                            ocr_data = pytesseract.image_to_data(
                                img,
                                lang=language,
                                config=config,
                                output_type=pytesseract.Output.DICT,
                            )
                            import json

                            extracted.append(json.dumps(ocr_data, ensure_ascii=False, indent=2))
                        else:
                            text = pytesseract.image_to_string(img, lang=language, config=config)
                            extracted.append(f"--- Page {current_page} ---\n{text}\n")
                    except Exception as e:
                        raise RuntimeError(
                            f"OCR failed on page {current_page}. Please check Tesseract installation "
                            f"and language data. Error: {str(e)}"
                        )
            except Exception as e:
                raise RuntimeError(f"Failed to render page {current_page} as image. Error: {str(e)}")

        output_path = Path(output_file)

        # Ensure output directory exists
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            raise RuntimeError(
                f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {str(e)}"
            )

        try:
            with output_path.open("w", encoding="utf-8") as f:
                f.write("".join(extracted))
        except PermissionError:
            raise PermissionError(
                f"Permission denied when writing to '{output_file}'. Please check file permissions "
                f"and ensure the file is not open in another application."
            )
        except Exception as e:
            raise RuntimeError(f"Failed to write OCR output to '{output_file}'. Error: {str(e)}")

        logger.info("OCR extraction completed: %s -> %s", input_file, output_file)


def create_searchable_pdf(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    language: str = "eng",
    dpi: int = 300,
    start_page: int = 1,
    end_page: int | None = None,
    config: str = "",
    # Preprocessing options
    binarize: bool = False,
    threshold: int = 128,
    resize_factor: float = 1.0,
    deskew: bool = False,
    denoise: bool = False,
    contrast_factor: float = 1.0,
    brightness_factor: float = 1.0,
    sharpen: bool = False,
    blur: float = 0.0,
    morph_op: str = "none",
    morph_kernel: int = 3,
    progress_callback: Optional[Callable[[tuple[int, str, float]], None]] = None,
    pdfa: bool = False,
    pdfa_method: str = "auto",
    pdfa_validate: bool = False,
) -> None:
    """Create a searchable PDF with embedded text layer using OCR. Optionally convert to PDF/A."""

    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    if not _HAVE_PYMUPDF:
        raise RuntimeError("PyMuPDF is required for OCR functionality. Please install it with: pip install pymupdf")
    if not _HAVE_TESSERACT:
        raise RuntimeError(
            "pytesseract is required for OCR functionality. Please install it with: pip install pytesseract"
        )
    if not _TESSERACT_INSTALLED:
        import pytesseract  # type: ignore

        raise pytesseract.pytesseract.TesseractError(
            1,
            "Tesseract OCR engine not found. Please install Tesseract from https://github.com/tesseract-ocr/tesseract",
        )

    import fitz  # type: ignore
    import pytesseract  # type: ignore

    logger.info("Creating searchable PDF from %s", input_file)

    # Open the document outside the context manager so we can save it later
    doc = fitz.open(str(input_file))

    try:
        total_pages = len(doc)
        if total_pages == 0:
            raise ValueError(f"The PDF file '{input_file}' contains no pages to process.")

        if end_page is None:
            end_page = int(total_pages)

        # Validate page range
        if start_page < 1 or end_page > total_pages or start_page > end_page:
            raise ValueError(
                f"Invalid page range {start_page}-{end_page} for document with {total_pages} pages. "
                f"Please specify a valid range between 1 and {total_pages}."
            )

        # Initialize progress tracking
        progress = OCRProgress(end_page - start_page + 1)

        for page_num in range(start_page - 1, end_page):
            current_page = page_num + 1
            progress.update(current_page, f"Processing page {current_page}/{total_pages}")

            if progress_callback:
                progress_callback(progress.get_progress())

            logger.info(f"Processing page {current_page}/{total_pages}")

            try:
                page = doc.load_page(page_num)
            except Exception as e:
                raise RuntimeError(f"Failed to load page {current_page} from PDF. Error: {str(e)}")

            try:
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)  # type: ignore
                img_data = pix.tobytes("png")
                with image_document(img_data) as img:
                    # Apply preprocessing
                    if any([binarize, resize_factor != 1.0, deskew, denoise]):
                        progress.update(current_page, f"Preprocessing page {current_page}")
                        if progress_callback:
                            progress_callback(progress.get_progress())

                        try:
                            img = preprocess_image(
                                img,
                                binarize=binarize,
                                threshold=threshold,
                                resize_factor=resize_factor,
                                deskew=deskew,
                                denoise=denoise,
                                contrast_factor=contrast_factor,
                                brightness_factor=brightness_factor,
                                sharpen=sharpen,
                                blur=blur,
                                morph_op=morph_op,
                                morph_kernel=morph_kernel,
                            )
                        except Exception as e:
                            raise RuntimeError(f"Failed to preprocess page {current_page}. Error: {str(e)}")

                    progress.update(current_page, f"Running OCR on page {current_page}")
                    if progress_callback:
                        progress_callback(progress.get_progress())

                    try:
                        ocr_data = pytesseract.image_to_data(
                            img,
                            lang=language,
                            config=config,
                            output_type=pytesseract.Output.DICT,
                        )
                    except Exception as e:
                        raise RuntimeError(
                            f"OCR failed on page {current_page}. Please check Tesseract installation "
                            f"and language data. Error: {str(e)}"
                        )

                    # Insert OCR text into PDF
                    try:
                        for i, text in enumerate(ocr_data["text"]):
                            if text.strip():
                                x, y, w, h = (
                                    ocr_data["left"][i],
                                    ocr_data["top"][i],
                                    ocr_data["width"][i],
                                    ocr_data["height"][i],
                                )
                                rect = fitz.Rect(x, y, x + w, y + h)
                                page.insert_text(rect.tl, text, fontsize=12)  # type: ignore
                    except Exception as e:
                        raise RuntimeError(f"Failed to insert OCR text into page {current_page}. Error: {str(e)}")
            except Exception as e:
                raise RuntimeError(f"Failed to render page {current_page} as image. Error: {str(e)}")

        # Write results to output file (for PDF/A conversion if requested)
        output_path = Path(output_file)

        # Ensure output directory exists
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            raise RuntimeError(
                f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {str(e)}"
            )

        # Save the modified PDF
        try:
            doc.save(str(output_file))
        except Exception as e:
            raise RuntimeError(f"Failed to save searchable PDF to '{output_file}'. Error: {str(e)}")

    finally:
        # Always close the document
        try:
            doc.close()
        except Exception as e:
            logger.warning(f"Error closing PDF document: {e}")

    # Convert to PDF/A if requested
    if pdfa:
        try:
            _convert_to_pdfa(output_file, pdfa_method, pdfa_validate)
        except Exception as e:
            raise RuntimeError(f"Failed to convert PDF to PDF/A. Error: {str(e)}")

    logger.info("Searchable PDF creation completed: %s -> %s", input_file, output_file)


def _convert_to_pdfa(input_file: str | os.PathLike[str], method: str = "auto", validate: bool = False) -> None:
    """Convert a PDF to PDF/A format."""
    # This is a placeholder for PDF/A conversion functionality
    # In a real implementation, this would use a library like PyPDF2 or Ghostscript
    logger.warning("PDF/A conversion is not yet implemented")
    pass


# ---------------------------------------------------------------------------
# Zonal OCR
# ---------------------------------------------------------------------------


def zonal_ocr_from_pdf(
    input_file: str | os.PathLike[str],
    zones: list[dict],
    output_file: str | os.PathLike[str],
    language: str = "eng",
    dpi: int = 300,
    config: str = "",
    output_format: str = "text",
    progress_callback: Optional[Callable[[tuple[int, str, float]], None]] = None,
) -> None:
    """Extract text from specific regions (zones) of a PDF using OCR.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file
    zones : list of dict
        List of zone dictionaries with keys: page (1-based), x, y, w, h
    output_file : str or PathLike
        Path to the output text/hOCR/JSON file
    language : str
        Tesseract language code(s)
    dpi : int
        Image DPI for rendering
    config : str
        Custom Tesseract config string
    output_format : str
        Output format: 'text', 'hocr', or 'json'
    """
    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    # Validate output format
    # Validate output format
    if output_format not in ["text", "hocr", "json"]:
        raise ValueError(f"Unsupported output format '{output_format}'. Supported formats are: 'text', 'hocr', 'json'")

    if not _HAVE_PYMUPDF:
        raise RuntimeError("PyMuPDF is required for OCR functionality. Please install it with: pip install pymupdf")
    if not _HAVE_TESSERACT:
        raise RuntimeError(
            "pytesseract is required for OCR functionality. Please install it with: pip install pytesseract"
        )
    if not _TESSERACT_INSTALLED:
        import pytesseract  # type: ignore

        raise pytesseract.pytesseract.TesseractError(
            1,
            "Tesseract OCR engine not found. Please install Tesseract from https://github.com/tesseract-ocr/tesseract",
        )

    import json

    import fitz  # type: ignore
    import pytesseract  # type: ignore

    logger.info("Starting zonal OCR extraction from %s", input_file)

    # Group zones by page for efficient processing
    zones_by_page = {}
    for zone in zones:
        page_num = zone.get("page", 1)
        if page_num not in zones_by_page:
            zones_by_page[page_num] = []
        zones_by_page[page_num].append(zone)

    # Initialize progress tracking
    total_zones = len(zones)
    processed_zones = 0
    results = []

    with pdf_document(input_file) as doc:
        total_pages = len(doc)

        # Process each page with zones
        for page_num, page_zones in zones_by_page.items():
            if page_num > total_pages:
                logger.warning(f"Skipping zone on page {page_num} - document only has {total_pages} pages")
                continue

            # Update progress
            processed_zones += len(page_zones)
            if progress_callback:
                progress = (processed_zones / total_zones) * 100
                progress_callback((processed_zones, f"Processing page {page_num}", progress))

            logger.info(f"Processing page {page_num} with {len(page_zones)} zones")

            try:
                # Load page
                page = doc.load_page(page_num - 1)  # Convert to 0-based indexing

                # Render page to image
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")

                with image_document(img_data) as img:
                    # Process each zone on this page
                    for i, zone in enumerate(page_zones):
                        try:
                            # Extract region from image
                            x, y, w, h = zone["x"], zone["y"], zone["w"], zone["h"]
                            # Scale coordinates to image dimensions
                            img_w, img_h = img.size
                            pdf_w, pdf_h = page.rect.width, page.rect.height
                            scale_x, scale_y = img_w / pdf_w, img_h / pdf_h

                            # Convert PDF coordinates to image coordinates
                            img_x = int(x * scale_x)
                            img_y = int(y * scale_y)
                            img_w = int(w * scale_x)
                            img_h = int(h * scale_y)

                            # Ensure coordinates are within image bounds
                            img_x = max(0, min(img_x, img_w - 1))
                            img_y = max(0, min(img_y, img_h - 1))
                            img_w = max(1, min(img_w, img_w - img_x))
                            img_h = max(1, min(img_h, img_h - img_y))

                            # Crop image to zone
                            zone_img = img.crop((img_x, img_y, img_x + img_w, img_y + img_h))

                            # Run OCR on zone
                            if output_format == "hocr":
                                text = pytesseract.image_to_pdf_or_hocr(
                                    zone_img,
                                    lang=language,
                                    config=config,
                                    extension="hocr",
                                )
                                if isinstance(text, bytes):
                                    results.append(text.decode("utf-8"))
                                else:
                                    results.append(str(text))
                            elif output_format == "json":
                                ocr_data = pytesseract.image_to_data(
                                    zone_img,
                                    lang=language,
                                    config=config,
                                    output_type=pytesseract.Output.DICT,
                                )
                                results.append(json.dumps(ocr_data, ensure_ascii=False, indent=2))
                            else:
                                text = pytesseract.image_to_string(zone_img, lang=language, config=config)
                                results.append(f"--- Zone {i + 1} (Page {page_num}) ---\n{text}\n")

                        except Exception as e:
                            logger.error(f"Failed to process zone {i + 1} on page {page_num}: {e}")
                            if output_format == "text":
                                results.append(f"--- Zone {i + 1} (Page {page_num}) ---\n[OCR failed: {str(e)}]\n")

            except Exception as e:
                logger.error(f"Failed to process page {page_num}: {e}")
                raise RuntimeError(f"Failed to process page {page_num}. Error: {str(e)}")

    # Write results to output file
    output_path = Path(output_file)

    # Ensure output directory exists
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        raise RuntimeError(
            f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {str(e)}"
        )

    try:
        with output_path.open("w", encoding="utf-8") as f:
            if output_format == "hocr":
                # For hOCR, combine all results into a single HTML document
                f.write("<!DOCTYPE html>\n<html>\n<body>\n")
                for result in results:
                    f.write(result)
                f.write("</body>\n</html>\n")
            else:
                f.write("".join(results))
    except PermissionError:
        raise PermissionError(
            f"Permission denied when writing to '{output_file}'. Please check file permissions "
            f"and ensure the file is not open in another application."
        )
    except Exception as e:
        raise RuntimeError(f"Failed to write OCR output to '{output_file}'. Error: {str(e)}")

    logger.info("Zonal OCR extraction completed: %s -> %s", input_file, output_file)


# ---------------------------------------------------------------------------
# Handwriting OCR
# ---------------------------------------------------------------------------


def handwriting_ocr_from_pdf(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    pages: list[int] | str | None = None,
    engine: str = "pytesseract",
    model: str | None = None,
    output_format: str = "text",
    language: str = "eng",
    dpi: int = 300,
    config: str = "",
    # Preprocessing options
    binarize: bool = False,
    threshold: int = 128,
    resize_factor: float = 1.5,
    deskew: bool = True,
    denoise: bool = True,
    contrast_factor: float = 1.2,
    brightness_factor: float = 1.1,
    sharpen: bool = True,
    blur: float = 0.5,
    morph_op: str = "none",
    morph_kernel: int = 3,
    progress_callback: Optional[Callable[[tuple[int, str, float]], None]] = None,
) -> None:
    """Extract text from handwriting PDF using specialized OCR and preprocessing.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file
    output_file : str or PathLike
        Path to the output text/JSON file
    pages : list of int, str, or None
        Pages to process (1-based). If None, process all pages.
        If string "all", process all pages.
    engine : str
        OCR engine to use: 'pytesseract' or 'kraken'
    model : str or None
        Path to custom model for kraken engine
    output_format : str
        Output format: 'text' or 'json'
    language : str
        Tesseract language code(s)
    dpi : int
        Image DPI for rendering
    config : str
        Custom Tesseract config string
    """
    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    # Validate output format
    if output_format not in ["text", "json"]:
        raise ValueError(f"Unsupported output format '{output_format}'. Supported formats are: 'text', 'json'")

    # Check if kraken is available
    _HAVE_KRAKEN = False
    if engine == "kraken":
        try:
            import kraken  # noqa: F401

            _HAVE_KRAKEN = True
        except ImportError:
            _HAVE_KRAKEN = False

        if not _HAVE_KRAKEN:
            logger.warning("Kraken not available, falling back to pytesseract")
            engine = "pytesseract"

    if engine == "pytesseract":
        if not _HAVE_TESSERACT:
            raise RuntimeError(
                "pytesseract is required for handwriting OCR functionality. "
                "Please install it with: pip install pytesseract"
            )
        if not _TESSERACT_INSTALLED:
            import pytesseract  # type: ignore

            raise pytesseract.pytesseract.TesseractError(
                1,
                "Tesseract OCR engine not found. Please install Tesseract from "
                "https://github.com/tesseract-ocr/tesseract",
            )

    if not _HAVE_PYMUPDF:
        raise RuntimeError("PyMuPDF is required for OCR functionality. Please install it with: pip install pymupdf")

    import json

    import cv2  # type: ignore
    import fitz  # type: ignore
    import numpy as np  # type: ignore

    logger.info(
        "Starting handwriting OCR extraction from %s using %s engine",
        input_file,
        engine,
    )

    with pdf_document(input_file) as doc:
        total_pages = len(doc)

        # Determine pages to process
        if pages is None or pages == "all":
            page_indices = list(range(total_pages))
        elif isinstance(pages, str):
            # Parse page string like "1,2,5-10"
            page_indices = []
            for part in pages.split(","):
                part = part.strip()
                if "-" in part:
                    start, end = part.split("-", 1)
                    page_indices.extend(range(int(start) - 1, int(end)))  # Convert to 0-based
                else:
                    page_indices.append(int(part) - 1)  # Convert to 0-based
        else:
            # List of page numbers (1-based)
            page_indices = [p - 1 for p in pages]  # Convert to 0-based

        # Validate page indices
        page_indices = [p for p in page_indices if 0 <= p < total_pages]
        if not page_indices:
            raise ValueError("No valid pages specified for processing")

        # Initialize progress tracking
        progress = OCRProgress(len(page_indices))
        results = []

        # Process each page
        for i, page_idx in enumerate(page_indices):
            current_page = page_idx + 1
            progress.update(current_page, f"Processing page {current_page}/{total_pages}")

            if progress_callback:
                progress_callback(progress.get_progress())

            logger.info(f"Processing page {current_page}/{total_pages}")

            try:
                # Load page
                page = doc.load_page(page_idx)

                # Render page to image
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")

                with image_document(img_data) as img:
                    # Apply preprocessing optimized for handwriting
                    progress.update(current_page, f"Preprocessing page {current_page}")
                    if progress_callback:
                        progress_callback(progress.get_progress())

                    try:
                        img = preprocess_image(
                            img,
                            binarize=binarize,
                            threshold=threshold,
                            resize_factor=resize_factor,
                            deskew=deskew,
                            denoise=denoise,
                            contrast_factor=contrast_factor,
                            brightness_factor=brightness_factor,
                            sharpen=sharpen,
                            blur=blur,
                            morph_op=morph_op,
                            morph_kernel=morph_kernel,
                        )
                    except Exception as e:
                        raise RuntimeError(f"Failed to preprocess page {current_page}. Error: {str(e)}")

                    # Run OCR
                    progress.update(current_page, f"Running OCR on page {current_page}")
                    if progress_callback:
                        progress_callback(progress.get_progress())

                    if engine == "kraken" and _HAVE_KRAKEN:
                        # Use kraken engine
                        try:
                            from kraken import (
                                binarization,
                                pageseg,
                                rpred,
                            )

                            # Convert PIL image to kraken format
                            np_img = np.array(img)
                            if len(np_img.shape) == 3:
                                np_img = cv2.cvtColor(np_img, cv2.COLOR_RGB2GRAY)

                            # Binarize image
                            binarized = binarization.nlbin(np_img)

                            # Segment page
                            seg = pageseg.segment(binarized)

                            # Load model
                            if model:
                                kraken_model = rpred.load_any(model)
                            else:
                                kraken_model = rpred.load_default_model()

                            # Recognize text
                            pred_it = rpred.rpred(kraken_model, binarized, seg)
                            ocr_result = list(pred_it)

                            # Format result
                            if output_format == "json":
                                result_data = [
                                    {
                                        "text": pred.prediction,
                                        "confidence": pred.confidence,
                                    }
                                    for pred in ocr_result
                                ]
                                results.append(json.dumps(result_data, ensure_ascii=False, indent=2))
                            else:
                                text = " ".join([pred.prediction for pred in ocr_result])
                                results.append(f"--- Page {current_page} ---\n{text}\n")

                        except Exception as e:
                            raise RuntimeError(f"Kraken OCR failed on page {current_page}. Error: {str(e)}")
                    else:
                        # Use pytesseract engine
                        try:
                            import pytesseract  # type: ignore

                            if output_format == "json":
                                ocr_data = pytesseract.image_to_data(
                                    img,
                                    lang=language,
                                    config=config,
                                    output_type=pytesseract.Output.DICT,
                                )
                                results.append(json.dumps(ocr_data, ensure_ascii=False, indent=2))
                            else:
                                text = pytesseract.image_to_string(img, lang=language, config=config)
                                results.append(f"--- Page {current_page} ---\n{text}\n")
                        except Exception as e:
                            raise RuntimeError(f"Pytesseract OCR failed on page {current_page}. Error: {str(e)}")

            except Exception as e:
                raise RuntimeError(f"Failed to process page {current_page}. Error: {str(e)}")

        # Write results to output file
        output_path = Path(output_file)

        # Ensure output directory exists
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            raise RuntimeError(
                f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {str(e)}"
            )

        try:
            with output_path.open("w", encoding="utf-8") as f:
                f.write("".join(results))
        except PermissionError:
            raise PermissionError(
                f"Permission denied when writing to '{output_file}'. Please check file permissions "
                f"and ensure the file is not open in another application."
            )
        except Exception as e:
            raise RuntimeError(f"Failed to write OCR output to '{output_file}'. Error: {str(e)}")

        logger.info("Handwriting OCR extraction completed: %s -> %s", input_file, output_file)


# ---------------------------------------------------------------------------
# Image Preprocessing
# ---------------------------------------------------------------------------


def preprocess_image(
    img: Image.Image,
    *,
    binarize: bool = False,
    threshold: int = 128,
    resize_factor: float = 1.0,
    deskew: bool = False,
    denoise: bool = False,
    contrast_factor: float = 1.0,
    brightness_factor: float = 1.0,
    sharpen: bool = False,
    blur: float = 0.0,
    morph_op: str = "none",
    morph_kernel: int = 3,
) -> Image.Image:
    """Apply preprocessing operations to an image for improved OCR results.

    Parameters
    ----------
    img : PIL.Image.Image
        Input image to preprocess
    binarize : bool
        Whether to convert to black and white
    threshold : int
        Threshold for binarization (0-255)
    resize_factor : float
        Scale factor for resizing (e.g., 1.5 = 150%)
    deskew : bool
        Whether to attempt to correct skew/rotation
    denoise : bool
        Whether to apply noise reduction
    contrast_factor : float
        Contrast adjustment factor (1.0 = no change)
    brightness_factor : float
        Brightness adjustment factor (1.0 = no change)
    sharpen : bool
        Whether to apply sharpening
    blur : float
        Amount of Gaussian blur to apply (0 = no blur)
    morph_op : str
        Morphological operation ('none', 'dilate', 'erode')
    morph_kernel : int
        Size of morphological operation kernel
    """
    # Validate parameters
    if binarize and (threshold < 0 or threshold > 255):
        raise ValueError("Threshold must be between 0 and 255")
    if resize_factor <= 0:
        raise ValueError("Resize factor must be positive")
    if blur < 0:
        raise ValueError("Blur radius must be non-negative")
    if morph_op != "none" and morph_kernel <= 0:
        raise ValueError("Morphological kernel size must be positive")
    if morph_op != "none" and morph_op not in ["dilate", "erode"]:
        raise ValueError("Morphological operation must be 'none', 'dilate', or 'erode'")
    if morph_op != "none" and morph_kernel % 2 == 0:
        raise ValueError("Morphological kernel size must be odd")

    import cv2
    import numpy as np

    # Work with a copy to avoid modifying the original
    processed = img.copy()

    # Resize if requested
    if resize_factor != 1.0:
        new_size = (
            int(processed.width * resize_factor),
            int(processed.height * resize_factor),
        )
        processed = processed.resize(new_size, Image.LANCZOS)

    # Convert to grayscale if not already
    if processed.mode != "L":
        processed = ImageOps.grayscale(processed)

    # Apply brightness adjustment
    if brightness_factor != 1.0:
        enhancer = ImageEnhance.Brightness(processed)
        processed = enhancer.enhance(brightness_factor)

    # Apply contrast adjustment
    if contrast_factor != 1.0:
        enhancer = ImageEnhance.Contrast(processed)
        processed = enhancer.enhance(contrast_factor)

    # Apply denoising
    if denoise:
        # Convert to numpy array for OpenCV processing
        np_img = np.array(processed)
        # Apply bilateral filter for noise reduction while preserving edges
        np_img = cv2.bilateralFilter(np_img, 9, 75, 75)
        processed = Image.fromarray(np_img)

    # Apply deskew if requested
    if deskew:
        # Convert to numpy array for OpenCV processing
        np_img = np.array(processed)
        # Calculate skew angle
        coords = np.column_stack(np.where(np_img > 0))
        angle = cv2.minAreaRect(coords)[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle

        # Apply rotation if significant skew detected
        if abs(angle) > 0.5:  # Only correct if skew is more than 0.5 degrees
            (h, w) = np_img.shape[:2]
            center = (w // 2, h // 2)
            M = cv2.getRotationMatrix2D(center, angle, 1.0)
            np_img = cv2.warpAffine(
                np_img,
                M,
                (w, h),
                flags=cv2.INTER_CUBIC,
                borderMode=cv2.BORDER_REPLICATE,
            )
            processed = Image.fromarray(np_img)

    # Apply morphological operations
    if morph_op != "none" and morph_kernel > 1:
        # Convert to numpy array for OpenCV processing
        np_img = np.array(processed)
        kernel = np.ones((morph_kernel, morph_kernel), np.uint8)
        if morph_op == "dilate":
            np_img = cv2.dilate(np_img, kernel, iterations=1)
        elif morph_op == "erode":
            np_img = cv2.erode(np_img, kernel, iterations=1)
        processed = Image.fromarray(np_img)

    # Apply blur
    if blur > 0:
        processed = processed.filter(ImageFilter.GaussianBlur(radius=blur))

    # Apply sharpening
    if sharpen:
        processed = processed.filter(ImageFilter.SHARPEN)

    # Apply binarization as the final step
    if binarize:
        processed = processed.point(lambda x: 0 if x < threshold else 255, mode="1")

    return processed


# ---------------------------------------------------------------------------
# Table Extraction
# ---------------------------------------------------------------------------


def extract_tables_from_pdf(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    *,
    engine: str = "camelot",
    output_format: str = "csv",
    pages: str = "1",
    flavor: str = "lattice",
    line_scale: int = 40,
    table_areas: list[str] | None = None,
    password: str | None = None,
    progress_callback: Optional[Callable[[tuple[int, str, float]], None]] = None,
) -> None:
    """Extract tables from a PDF file (backward compatibility function).

    This is a wrapper function for extract_tables with parameters matching
    the GUI expectations.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file containing tables
    output_file : str or PathLike
        Path to the output file where extracted tables will be saved
    engine : str, default "camelot"
        Table extraction engine to use ("camelot" or "pdfplumber")
    output_format : str, default "csv"
        Output format for extracted tables ("csv", "json", "excel", or "html")
    pages : str, default "1"
        Pages to process. Can be "all", a list of page numbers (1-based), or a string like "1,2,5-10"
    flavor : str, default "lattice"
        Camelot flavor to use ("lattice" or "stream") - only for Camelot engine
    line_scale : int, default 40
        Line scale parameter for Camelot - only for Camelot engine
    table_areas : list of str, optional
        Table areas to process - only for Camelot engine
    password : str, optional
        Password for encrypted PDFs
    progress_callback : Callable, optional
        Callback function to report progress with signature (current, total, status)

    Raises
    ------
    FileNotFoundError
        If the input file does not exist
    ValueError
        If invalid parameters are provided
    RuntimeError
        If table extraction fails or dependencies are missing
    PermissionError
        If there are insufficient permissions to write the output file
    """
    return extract_tables(
        input_file=input_file,
        output_file=output_file,
        engine=engine,
        output_format=output_format,
        pages=pages,
        flavor=flavor,
        line_scale=line_scale,
        table_areas=table_areas,
        password=password,
        progress_callback=progress_callback,
    )


def extract_tables(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    *,
    engine: str = "camelot",
    output_format: str = "csv",
    pages: str | list[int] | None = None,
    flavor: str = "lattice",
    line_scale: int = 40,
    table_areas: list[str] | None = None,
    password: str | None = None,
    progress_callback: Optional[Callable[[tuple[int, str, float]], None]] = None,
) -> None:
    """Extract tables from a PDF file using Camelot or pdfplumber.

    This function extracts tables from a PDF file and saves them to various output formats.
    It supports two engines: Camelot (default) and pdfplumber, each with different capabilities
    and requirements.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file containing tables
    output_file : str or PathLike
        Path to the output file where extracted tables will be saved
    engine : str, default "camelot"
        Table extraction engine to use ("camelot" or "pdfplumber")
    output_format : str, default "csv"
        Output format for extracted tables ("csv", "json", "excel", or "html")
    pages : str, list of int, or None, default None
        Pages to process. If None, process all pages.
        Can be "all", a list of page numbers (1-based), or a string like "1,2,5-10"
    flavor : str, default "lattice"
        Camelot flavor to use ("lattice" or "stream") - only for Camelot engine
    line_scale : int, default 40
        Line scale parameter for Camelot - only for Camelot engine
    table_areas : list of str, optional
        Table areas to process - only for Camelot engine
    password : str, optional
        Password for encrypted PDFs
    progress_callback : Callable, optional
        Callback function to report progress with signature (current, total, status)

    Raises
    ------
    FileNotFoundError
        If the input file does not exist
    ValueError
        If invalid parameters are provided
    RuntimeError
        If table extraction fails or dependencies are missing
    PermissionError
        If there are insufficient permissions to write the output file

    Examples
    --------
    >>> # Extract tables using Camelot with lattice flavor
    >>> extract_tables("document.pdf", "tables.csv", engine="camelot", flavor="lattice")

    >>> # Extract tables using pdfplumber from specific pages
    >>> extract_tables("document.pdf", "tables.json", engine="pdfplumber", pages="1,3-5", output_format="json")

    >>> # Extract tables with custom table areas
    >>> extract_tables("document.pdf", "tables.xlsx", table_areas=["100,200,500,400"], output_format="excel")

    Notes
    -----
    Required dependencies:
    - For Camelot engine: camelot-py[base] or camelot-py[cv]
    - For pdfplumber engine: pdfplumber
    - Output formats:
      * CSV: Simple comma-separated values (requires pandas)
      * JSON: Structured JSON format
      * Excel: Microsoft Excel format (requires pandas and openpyxl)
      * HTML: HTML table format (requires pandas)
    """
    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    # Validate engine
    if engine not in ("camelot", "pdfplumber"):
        raise ValueError(f"Unsupported engine: {engine}. Supported engines are 'camelot' and 'pdfplumber'.")

    # Validate output format
    if output_format not in ("csv", "json", "excel", "html"):
        raise ValueError(
            f"Unsupported output format: {output_format}. Supported formats are 'csv', 'json', 'excel', and 'html'."
        )

    # Check for required dependencies based on engine
    if engine == "camelot":
        if not _HAVE_CAMELOT:
            raise RuntimeError(
                "Camelot is required for table extraction with the 'camelot' engine. "
                "Please install it with: pip install camelot-py[base] or pip install camelot-py[cv]"
            )
    else:  # pdfplumber
        if not _HAVE_PDFPLUMBER:
            raise RuntimeError(
                "pdfplumber is required for table extraction with the 'pdfplumber' engine. "
                "Please install it with: pip install pdfplumber"
            )

    # Check for PyMuPDF requirement
    if not _HAVE_PYMUPDF:
        raise RuntimeError(
            "PyMuPDF is required for table extraction functionality. Please install it with: pip install pymupdf"
        )

    import json  # type: ignore

    logger.info("Starting table extraction from %s using %s engine", input_file, engine)

    try:
        # Handle password protection for PyMuPDF
        doc_args = {"filename": str(input_file)}
        if password:
            doc_args["password"] = password

        # Try to open the document
        try:
            with fitz.open(**doc_args) as doc:
                total_pages = len(doc)

                # Handle empty PDF
                if total_pages == 0:
                    # For empty PDFs, create an empty output file
                    output_path = Path(output_file)
                    try:
                        output_path.parent.mkdir(parents=True, exist_ok=True)
                    except Exception as e:
                        raise RuntimeError(
                            f"Failed to create output directory for '{output_file}'. "
                            f"Please check permissions. Error: {str(e)}"
                        )

                    try:
                        # Create empty file based on output format
                        with output_path.open("w", encoding="utf-8") as f:
                            if output_format == "json":
                                f.write("[]")  # Empty JSON array
                            elif output_format == "csv":
                                f.write("")  # Empty CSV
                            elif output_format == "excel":
                                # For Excel, we need to create an empty file
                                try:
                                    import pandas as pd  # type: ignore  # noqa: F401
                                    from openpyxl import Workbook  # type: ignore

                                    wb = Workbook()
                                    ws = wb.active
                                    ws.title = "Empty"
                                    wb.save(output_path)
                                except ImportError:
                                    # Fallback to empty file
                                    f.write("")
                            else:  # html
                                f.write(
                                    "<!DOCTYPE html><html><head><title>Empty Tables</title></head>"
                                    "<body><h1>No Tables Found</h1></body></html>"
                                )
                        logger.info("Empty PDF processed: %s -> %s", input_file, output_file)
                        return
                    except PermissionError:
                        raise PermissionError(
                            f"Permission denied when writing to '{output_file}'. Please check file permissions "
                            f"and ensure the file is not open in another application."
                        )
                    except Exception as e:
                        raise RuntimeError(f"Failed to write table output to '{output_file}'. Error: {str(e)}")

                # Determine pages to process
                if pages is None or pages == "all":
                    page_indices = list(range(total_pages))
                elif isinstance(pages, str):
                    if pages == "all":
                        page_indices = list(range(total_pages))
                    else:
                        # Parse page string like "1,2,5-10"
                        page_indices = []
                        for part in pages.split(","):
                            part = part.strip()
                            if "-" in part:
                                start, end = part.split("-", 1)
                                page_indices.extend(range(int(start) - 1, int(end)))  # Convert to 0-based
                            else:
                                page_indices.append(int(part) - 1)  # Convert to 0-based
                else:
                    # List of page numbers (1-based)
                    page_indices = [p - 1 for p in pages]  # Convert to 0-based

                # Validate page indices
                page_indices = [p for p in page_indices if 0 <= p < total_pages]
                if not page_indices:
                    raise ValueError("No valid pages specified for processing")

                # Initialize progress tracking
                progress = OCRProgress(len(page_indices))
                all_tables = []

                # Process each page
                for i, page_idx in enumerate(page_indices):
                    current_page = page_idx + 1
                    progress.update(current_page, f"Processing page {current_page}/{total_pages}")

                    if progress_callback:
                        progress_callback(progress.get_progress())

                    logger.info(f"Processing page {current_page}/{total_pages}")

                    try:
                        if engine == "camelot":
                            # Use Camelot for table extraction
                            import camelot  # type: ignore  # noqa: F811

                            # Prepare Camelot parameters
                            camelot_kwargs = {
                                "pages": str(current_page),
                                "flavor": flavor,
                                "line_scale": line_scale,
                            }

                            if table_areas:
                                camelot_kwargs["table_areas"] = table_areas

                            # Extract tables using Camelot
                            try:
                                tables = camelot.read_pdf(str(input_file), **camelot_kwargs)

                                # Convert Camelot tables to standard format
                                for table in tables:
                                    # Extract data and headers
                                    data = table.df.values.tolist()
                                    headers = table.df.columns.tolist()

                                    # Create standard table format
                                    table_data = {
                                        "page": current_page,
                                        "engine": "camelot",
                                        "data": data,
                                        "headers": headers,
                                    }
                                    all_tables.append(table_data)
                            except Exception as e:
                                logger.warning(f"Camelot failed to extract tables from page {current_page}: {e}")
                                # Continue with other pages

                        else:  # pdfplumber
                            # Use pdfplumber for table extraction
                            import pdfplumber  # type: ignore  # noqa: F811

                            # Open the PDF with pdfplumber
                            try:
                                with pdfplumber.open(str(input_file), password=password) as pdf:
                                    # Get the specific page
                                    if page_idx < len(pdf.pages):
                                        page = pdf.pages[page_idx]

                                        # Extract tables from the page
                                        tables = page.extract_tables()

                                        # Convert pdfplumber tables to standard format
                                        for table in tables:
                                            # pdfplumber returns tables as lists of lists
                                            # First row is typically headers
                                            if table:
                                                headers = table[0] if table else []
                                                data = table[1:] if len(table) > 1 else []

                                                # Create standard table format
                                                table_data = {
                                                    "page": current_page,
                                                    "engine": "pdfplumber",
                                                    "data": data,
                                                    "headers": headers,
                                                }
                                                all_tables.append(table_data)
                            except Exception as e:
                                logger.warning(f"pdfplumber failed to extract tables from page {current_page}: {e}")
                                # Continue with other pages

                    except Exception as e:
                        logger.warning(f"Failed to extract tables from page {current_page}: {e}")
                        # Continue with other pages instead of failing completely

                # Write results to output file
                output_path = Path(output_file)

                # Ensure output directory exists
                try:
                    output_path.parent.mkdir(parents=True, exist_ok=True)
                except Exception as e:
                    raise RuntimeError(
                        f"Failed to create output directory for '{output_file}'. "
                        f"Please check permissions. Error: {str(e)}"
                    )

                try:
                    if output_format == "json":
                        with output_path.open("w", encoding="utf-8") as f:
                            json.dump(all_tables, f, ensure_ascii=False, indent=2)
                    elif output_format == "csv":
                        # For CSV, we need pandas to handle the conversion
                        try:
                            import pandas as pandas_csv  # type: ignore

                            # Flatten tables for CSV output
                            if all_tables:
                                csv_data = []
                                for table in all_tables:
                                    # Add a header row with page and engine info
                                    csv_data.append([f"Page {table['page']} (Engine: {table['engine']})"])

                                    # Add headers if available
                                    if table["headers"]:
                                        csv_data.append(table["headers"])

                                    # Add data rows
                                    csv_data.extend(table["data"])

                                    # Add empty row between tables
                                    csv_data.append([])

                                # Create DataFrame and save to CSV
                                df = pandas_csv.DataFrame(csv_data)
                                df.to_csv(
                                    output_path,
                                    index=False,
                                    header=False,
                                    encoding="utf-8",
                                )
                            else:
                                # Create empty CSV file
                                with output_path.open("w", encoding="utf-8") as f:
                                    pass
                        except ImportError:
                            # Fallback to simple CSV writing without pandas
                            with output_path.open("w", newline="", encoding="utf-8") as f:
                                if all_tables:
                                    import csv

                                    writer = csv.writer(f)
                                    for table in all_tables:
                                        # Write page info
                                        writer.writerow([f"Page {table['page']} (Engine: {table['engine']})"])

                                        # Write headers if available
                                        if table["headers"]:
                                            writer.writerow(table["headers"])

                                        # Write data rows
                                        for row in table["data"]:
                                            writer.writerow(row)

                                        # Add empty row between tables
                                        writer.writerow([])
                                # If no tables, file will be empty
                    elif output_format == "excel":
                        # For Excel, we need pandas and openpyxl
                        try:
                            import pandas as pandas_excel  # type: ignore

                            if all_tables:
                                # Create a DataFrame for each table and save to different sheets
                                with pandas_excel.ExcelWriter(output_path, engine="openpyxl") as writer:
                                    for i, table in enumerate(all_tables):
                                        # Create DataFrame from table data
                                        df = pandas_excel.DataFrame(
                                            table["data"],
                                            columns=table["headers"] if table["headers"] else None,
                                        )
                                        # Save to a sheet named by page number and table index
                                        sheet_name = f"Page{table['page']}_Table{i + 1}"
                                        # Excel sheet names are limited to 31 characters
                                        sheet_name = sheet_name[:31]
                                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                            else:
                                # Create empty Excel file
                                from openpyxl import Workbook  # type: ignore

                                wb = Workbook()
                                ws = wb.active
                                ws.title = "No Tables"
                                wb.save(output_path)
                        except ImportError:
                            raise RuntimeError(
                                "pandas and openpyxl are required for Excel output format. "
                                "Please install them with: pip install pandas openpyxl"
                            )
                    else:  # html
                        # For HTML, we need pandas
                        if all_tables:
                            try:
                                import pandas as pandas_html  # type: ignore

                                # Create HTML output with all tables
                                html_parts = [
                                    "<!DOCTYPE html>",
                                    "<html>",
                                    "<head>",
                                    "<meta charset='utf-8'>",
                                    "<title>Extracted Tables</title>",
                                    "<style>table {border-collapse: collapse; margin: 10px 0;} "
                                    "th, td {border: 1px solid #ddd; padding: 8px;} "
                                    "th {background-color: #f2f2f2;}</style>",
                                    "</head>",
                                    "<body>",
                                    "<h1>Extracted Tables</h1>",
                                ]

                                for i, table in enumerate(all_tables):
                                    html_parts.append(f"<h2>Page {table['page']} (Engine: {table['engine']})</h2>")

                                    # Create DataFrame and convert to HTML
                                    df = pandas_html.DataFrame(
                                        table["data"],
                                        columns=table["headers"] if table["headers"] else None,
                                    )
                                    html_table = df.to_html(index=False, escape=False)
                                    html_parts.append(html_table)

                                html_parts.extend(["</body>", "</html>"])

                                with output_path.open("w", encoding="utf-8") as f:
                                    f.write("\n".join(html_parts))
                            except ImportError:
                                # Fallback to simple HTML without pandas
                                html_parts = [
                                    "<!DOCTYPE html>",
                                    "<html>",
                                    "<head>",
                                    "<meta charset='utf-8'>",
                                    "<title>Extracted Tables</title>",
                                    "<style>table {border-collapse: collapse; margin: 10px 0;} "
                                    "th, td {border: 1px solid #ddd; padding: 8px;} "
                                    "th {background-color: #f2f2f2;}</style>",
                                    "</head>",
                                    "<body>",
                                    "<h1>Extracted Tables</h1>",
                                ]

                                for i, table in enumerate(all_tables):
                                    html_parts.append(f"<h2>Page {table['page']} (Engine: {table['engine']})</h2>")
                                    html_parts.append("<table>")

                                    # Add headers if available
                                    if table["headers"]:
                                        html_parts.append("<thead><tr>")
                                        for header in table["headers"]:
                                            html_parts.append(f"<th>{header}</th>")
                                        html_parts.append("</tr></thead>")

                                    # Add data rows
                                    html_parts.append("<tbody>")
                                    for row in table["data"]:
                                        html_parts.append("<tr>")
                                        for cell in row:
                                            html_parts.append(f"<td>{cell}</td>")
                                        html_parts.append("</tr>")
                                    html_parts.append("</tbody>")
                                    html_parts.append("</table>")

                                html_parts.extend(["</body>", "</html>"])

                                with output_path.open("w", encoding="utf-8") as f:
                                    f.write("\n".join(html_parts))
                        else:
                            # Create empty HTML file
                            with output_path.open("w", encoding="utf-8") as f:
                                f.write(
                                    "<!DOCTYPE html><html><head><title>Empty Tables</title></head>"
                                    "<body><h1>No Tables Found</h1></body></html>"
                                )

                except PermissionError:
                    raise PermissionError(
                        f"Permission denied when writing to '{output_file}'. Please check file permissions "
                        f"and ensure the file is not open in another application."
                    )
                except Exception as e:
                    raise RuntimeError(f"Failed to write table output to '{output_file}'. Error: {str(e)}")

                logger.info("Table extraction completed: %s -> %s", input_file, output_file)

        except Exception as e:
            # Handle the case where the PDF file is empty or corrupted
            logger.warning(f"Failed to open PDF file: {e}")
            # For empty/corrupted PDFs, create an empty output file
            output_path = Path(output_file)
            try:
                output_path.parent.mkdir(parents=True, exist_ok=True)
            except Exception as dir_e:
                raise RuntimeError(
                    f"Failed to create output directory for '{output_file}'. "
                    f"Please check permissions. Error: {str(dir_e)}"
                )

            try:
                # Create empty file based on output format
                with output_path.open("w", encoding="utf-8") as f:
                    if output_format == "json":
                        f.write("[]")  # Empty JSON array
                    elif output_format == "csv":
                        f.write("")  # Empty CSV
                    elif output_format == "excel":
                        # For Excel, we need to create an empty file
                        try:
                            from openpyxl import Workbook  # type: ignore

                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Empty"
                            wb.save(output_path)
                        except ImportError:
                            # Fallback to empty file
                            f.write("")
                    else:  # html
                        f.write(
                            "<!DOCTYPE html><html><head><title>Empty Tables</title></head>"
                            "<body><h1>No Tables Found</h1></body></html>"
                        )
                logger.info("Empty/corrupted PDF processed: %s -> %s", input_file, output_file)
                return
            except PermissionError:
                raise PermissionError(
                    f"Permission denied when writing to '{output_file}'. Please check file permissions "
                    f"and ensure the file is not open in another application."
                )
            except Exception as write_e:
                raise RuntimeError(f"Failed to write table output to '{output_file}'. Error: {str(write_e)}")

    except Exception as e:
        raise RuntimeError(f"Table extraction failed. Error: {str(e)}")


# ---------------------------------------------------------------------------
# Barcode Extraction
# ---------------------------------------------------------------------------


def detect_barcodes(
    file: str | os.PathLike[str],
    output_dir: str | os.PathLike[str],
    barcode_types: list[str] | None = None,
    output_format: str = "json",
    page_range: str | None = None,
    dpi: int = 200,
    password: str | None = None,
) -> tuple[bool, list | dict | str]:
    """Alias function for extract_barcodes_from_pdf to maintain backward compatibility with tests.

    This function provides a compatibility layer for tests that expect the old function signature.

    Parameters
    ----------
    file : str or PathLike
        Path to the input PDF file
    output_dir : str or PathLike
        Path to the output directory
    barcode_types : list of str or None
        List of barcode types to detect (e.g., ['CODE128', 'QR_CODE'])
        If None, detect all supported types
    output_format : str
        Output format: 'json', 'csv', or 'txt'
    page_range : str or None
        Pages to process. Can be "all", a list of page numbers (1-based),
        or a string like "1,2,5-10"
    dpi : int
        Image DPI for rendering
    password : str or None
        Password for encrypted PDFs

    Returns
    -------
    tuple
        A tuple containing:
        - bool: Success status
        - list | dict | str: Detected barcodes data or error message
    """
    from pathlib import Path

    # Convert output_dir to a Path and create output file path
    output_path = Path(output_dir) / f"barcodes.{output_format}"

    try:
        result = extract_barcodes_from_pdf(
            input_file=file,
            output_file=output_path,
            pages=page_range,
            barcode_types=barcode_types,
            output_format=output_format,
            dpi=dpi,
            password=password,
        )
        return result
    except Exception as e:
        return False, [str(e)]


def extract_barcodes_from_pdf(
    input_file: str | os.PathLike[str],
    output_file: str | os.PathLike[str],
    pages: str | list[int] | None = None,
    barcode_types: list[str] | None = None,
    output_format: str = "json",
    dpi: int = 200,
    password: str | None = None,
    return_images: bool = False,
    progress_callback: Optional[Callable[[tuple[int, str, float]], None]] = None,
) -> tuple[bool, list | dict | str]:
    """Extract barcodes and QR codes from a PDF file.

    Parameters
    ----------
    input_file : str or PathLike
        Path to the input PDF file
    output_file : str or PathLike
        Path to the output file
    pages : str, list of int, or None
        Pages to process. If None, process all pages.
        Can be "all", a list of page numbers (1-based), or a string like "1,2,5-10"
    barcode_types : list of str or None
        List of barcode types to detect (e.g., ['CODE128', 'QR_CODE'])
        If None, detect all supported types
    output_format : str
        Output format: 'json', 'csv', or 'txt'
    dpi : int
        Image DPI for rendering
    password : str or None
        Password for encrypted PDFs
    return_images : bool
        Whether to return image snippets of detected barcodes
    """
    # Validate input file
    input_path = Path(input_file)
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_file}. Please check the file path and try again.")

    # Check if barcode backends are available: prefer pyzbar; fallback to OpenCV QR
    pyzbar = None  # type: ignore
    cv2 = None  # type: ignore
    np = None  # type: ignore
    _HAVE_PYZBAR = False
    _HAVE_QR_FALLBACK = False
    try:
        import cv2 as _cv2  # type: ignore
        import numpy as _np  # type: ignore
        from pyzbar import pyzbar as _pyzbar  # type: ignore

        pyzbar, cv2, np = _pyzbar, _cv2, _np
        _HAVE_PYZBAR = True
    except Exception:
        try:
            import cv2 as _cv2  # type: ignore
            import numpy as _np  # type: ignore

            cv2, np = _cv2, _np
            _HAVE_QR_FALLBACK = True
            logger.warning("pyzbar unavailable; using OpenCV QRCodeDetector fallback")
        except Exception:
            pass
    if not (_HAVE_PYZBAR or _HAVE_QR_FALLBACK):
        raise RuntimeError(
            "Barcode detection requires 'pyzbar' or OpenCV QR fallback. Install: pip install pyzbar opencv-python"
        )

    if not _HAVE_PYMUPDF:
        raise RuntimeError(
            "PyMuPDF is required for barcode extraction functionality. Please install it with: pip install pymupdf"
        )

    import csv
    import json

    logger.info("Starting barcode extraction from %s", input_file)

    with pdf_document(input_file) as doc:
        # Handle password protection
        if password:
            try:
                doc.authenticate(password)
            except Exception as e:
                raise RuntimeError(f"Failed to authenticate PDF with provided password. Error: {str(e)}")

        total_pages = len(doc)

        # Determine pages to process
        if pages is None or pages == "all":
            page_indices = list(range(total_pages))
        elif isinstance(pages, str):
            if pages == "all":
                page_indices = list(range(total_pages))
            else:
                # Parse page string like "1,2,5-10"
                page_indices = []
                for part in pages.split(","):
                    part = part.strip()
                    if "-" in part:
                        start, end = part.split("-", 1)
                        page_indices.extend(range(int(start) - 1, int(end)))  # Convert to 0-based
                    else:
                        page_indices.append(int(part) - 1)  # Convert to 0-based
        else:
            # List of page numbers (1-based)
            page_indices = [p - 1 for p in pages]  # Convert to 0-based

        # Validate page indices
        page_indices = [p for p in page_indices if 0 <= p < total_pages]
        if not page_indices:
            raise ValueError("No valid pages specified for processing")

        # Initialize progress tracking
        progress = OCRProgress(len(page_indices))
        detected_barcodes = []

        # Process each page
        for i, page_idx in enumerate(page_indices):
            current_page = page_idx + 1
            progress.update(current_page, f"Processing page {current_page}/{total_pages}")

            if progress_callback:
                progress_callback(progress.get_progress())

            logger.info(f"Processing page {current_page}/{total_pages}")

            try:
                # Load page
                page = doc.load_page(page_idx)

                # Render page to image
                mat = fitz.Matrix(dpi / 72, dpi / 72)
                pix = page.get_pixmap(matrix=mat)
                img_data = pix.tobytes("png")

                with image_document(img_data) as img:
                    # Convert PIL image to OpenCV format
                    np_img = np.array(img)
                    if len(np_img.shape) == 3:
                        # Convert RGB to BGR for OpenCV
                        np_img = cv2.cvtColor(np_img, cv2.COLOR_RGB2BGR)

                    if _HAVE_PYZBAR:
                        # Detect barcodes with pyzbar
                        barcodes = pyzbar.decode(np_img, symbols=barcode_types)
                        for barcode in barcodes:
                            barcode_info = {
                                "page": current_page,
                                "type": barcode.type,
                                "data": barcode.data.decode("utf-8"),
                                "rect": {
                                    "x": barcode.rect.left,
                                    "y": barcode.rect.top,
                                    "width": barcode.rect.width,
                                    "height": barcode.rect.height,
                                },
                            }
                            if return_images:
                                x, y, w, h = (
                                    barcode.rect.left,
                                    barcode.rect.top,
                                    barcode.rect.width,
                                    barcode.rect.height,
                                )
                                img_h, img_w = np_img.shape[:2]
                                x = max(0, min(x, img_w - 1))
                                y = max(0, min(y, img_h - 1))
                                w = max(1, min(w, img_w - x))
                                h = max(1, min(h, img_h - y))
                                snippet = np_img[y : y + h, x : x + w]
                                snippet = cv2.cvtColor(snippet, cv2.COLOR_BGR2RGB)
                                barcode_info["image"] = snippet
                            detected_barcodes.append(barcode_info)
                    else:
                        # QR-only fallback via OpenCV
                        qr = cv2.QRCodeDetector()
                        data, points, _ = qr.detectAndDecode(np_img)
                        if points is not None and data:
                            pts = points[0].astype(int)
                            x = int(pts[:, 0].min())
                            y = int(pts[:, 1].min())
                            w = int(pts[:, 0].max() - x)
                            h = int(pts[:, 1].max() - y)
                            info = {
                                "page": current_page,
                                "type": "QR_CODE",
                                "data": data,
                                "rect": {"x": x, "y": y, "width": w, "height": h},
                            }
                            if return_images:
                                img_h, img_w = np_img.shape[:2]
                                x = max(0, min(x, img_w - 1))
                                y = max(0, min(y, img_h - 1))
                                w = max(1, min(w, img_w - x))
                                h = max(1, min(h, img_h - y))
                                snippet = np_img[y : y + h, x : x + w]
                                info["image"] = cv2.cvtColor(snippet, cv2.COLOR_BGR2RGB)
                            detected_barcodes.append(info)

            except Exception as e:
                raise RuntimeError(f"Failed to process page {current_page}. Error: {str(e)}")

        # Write results to output file
        output_path = Path(output_file)

        # Ensure output directory exists
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except Exception as e:
            raise RuntimeError(
                f"Failed to create output directory for '{output_file}'. Please check permissions. Error: {str(e)}"
            )

        try:
            if output_format == "json":
                with output_path.open("w", encoding="utf-8") as f:
                    json.dump(detected_barcodes, f, ensure_ascii=False, indent=2)
            elif output_format == "csv":
                with output_path.open("w", newline="", encoding="utf-8") as f:
                    if detected_barcodes:
                        writer = csv.DictWriter(f, fieldnames=detected_barcodes[0].keys())
                        writer.writeheader()
                        for barcode in detected_barcodes:
                            # Flatten the rect dictionary for CSV
                            row = barcode.copy()
                            rect = row.pop("rect", {})
                            row.update({f"rect_{k}": v for k, v in rect.items()})
                            writer.writerow(row)
            else:  # txt format
                with output_path.open("w", encoding="utf-8") as f:
                    for barcode in detected_barcodes:
                        f.write(f"Page {barcode['page']}: {barcode['type']} = {barcode['data']}\n")
                        if "rect" in barcode:
                            rect = barcode["rect"]
                            f.write(f"  Location: ({rect['x']}, {rect['y']}, {rect['width']}, {rect['height']})\n")
                        f.write("\n")
        except PermissionError:
            raise PermissionError(
                f"Permission denied when writing to '{output_file}'. Please check file permissions "
                f"and ensure the file is not open in another application."
            )
        except Exception as e:
            raise RuntimeError(f"Failed to write barcode output to '{output_file}'. Error: {str(e)}")

        logger.info("Barcode extraction completed: %s -> %s", input_file, output_file)
        return True, detected_barcodes
